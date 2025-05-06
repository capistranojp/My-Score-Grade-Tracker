import os
import tkinter as tk
from tkinter import Frame, messagebox, ttk
import customtkinter as ctk
import openpyxl
from openpyxl.styles import Font

# App setup
mw = ctk.CTk()
mw.title("Score Tracker/Grade Tracker")
mw.geometry("1020x380")
mw.minsize(1020, 380)
mw.grid_columnconfigure(3, weight=1)
mw.grid_rowconfigure(0, weight=1)

filename = "student_data.xlsx"
passing = 75

def workb():
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        headers = ["Name", "Score 1", "Score 2", "Score 3", "Score 4", "Score 5", "Average", "Status"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
    return wb, ws

def save_student_record(name, scores):
    wb, ws = workb()
    average = round(sum(scores) / len(scores), 2)
    status = "Pass" if average >= passing else "Fail"
    data = [name] + scores + [average, status]
    ws.append(data)
    wb.save(filename)

def calculate_average(scores):
    return round(sum(scores) / len(scores), 2)

def display_all_records():
    if not os.path.exists(filename):
        return
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    records_list.delete(*records_list.get_children())
    for row in ws.iter_rows(min_row=2, values_only=True):
        records_list.insert("", "end", values=row)

def clear_entries():
    student_tbox.delete(0, ctk.END)
    for entry in score_entries:
        entry.delete(0, ctk.END)
    average_entry.configure(state="normal")
    average_entry.delete(0, ctk.END)
    average_entry.configure(state="readonly")

def calculate_average_event():
    try:
        scores = [float(entry.get()) for entry in score_entries]
        if any(score > 100 for score in scores):
            raise ValueError("Scores cannot be more than 100")
        avg = calculate_average(scores)
        average_entry.configure(state="normal")
        average_entry.delete(0, tk.END)
        average_entry.insert(0, str(avg))
        average_entry.configure(state="readonly")
    except ValueError:
        messagebox.showerror("Input Error", "Please enter valid numbers (0-100) for all scores.")

def save_record():
    name = student_tbox.get().strip()
    try:
        scores = [float(entry.get()) for entry in score_entries]
        if not name:
            raise ValueError("Name is required")
        if any(score > 100 for score in scores):
            raise ValueError("Scores cannot be more than 100")
        if passing > 100:
            raise ValueError("Passing score cannot be more than 100")

        save_student_record(name, scores)
        clear_entries()
        display_all_records()
        messagebox.showinfo("Saved", f"Record saved for {name}")
    except ValueError as e:
        messagebox.showerror("Error", str(e))

def clear_table():
    if not os.path.exists(filename):
        return
    confirm = messagebox.askyesno("Confirm", "Are you sure you want to clear all records?")
    if confirm:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ws.delete_rows(2, ws.max_row - 1)
        wb.save(filename)
        display_all_records()
        messagebox.showinfo("Cleared", "All records have been cleared.")

# Labels and entries (left side)
student_tbox_label = ctk.CTkLabel(mw, text="Student's Name:")
student_tbox_label.grid(row=0, column=0, sticky="W", padx=10, pady=5)
student_tbox = ctk.CTkEntry(mw)
student_tbox.grid(row=0, column=1, sticky="W", padx=10, pady=5)

score_labels = ["Score 1:", "Score 2:", "Score 3:", "Score 4:", "Score 5:"]
score_entries = []
for idx, label_text in enumerate(score_labels, start=1):
    score_label = ctk.CTkLabel(mw, text=label_text)
    score_label.grid(row=idx, column=0, sticky="W", padx=10, pady=5)
    entry = ctk.CTkEntry(mw)
    entry.grid(row=idx, column=1, sticky="W", padx=10, pady=5)
    score_entries.append(entry)

average_label = ctk.CTkLabel(mw, text="Average:")
average_label.grid(row=6, column=0, sticky="W", padx=10, pady=5)
average_entry = ctk.CTkEntry(mw, state="readonly")
average_entry.grid(row=6, column=1, sticky="W", padx=10, pady=5)

# Separator line
separator = ttk.Separator(mw, orient="vertical")
separator.grid(row=0, column=2, rowspan=8, sticky="ns", padx=10)

# Record display with scrollbars
records_frame = Frame(mw)
records_frame.grid(row=0, column=3, rowspan=7, padx=10, pady=5, sticky="nsew")
records_frame.grid_rowconfigure(0, weight=1)
records_frame.grid_columnconfigure(0, weight=1)

columns = ("Name", "Score 1", "Score 2", "Score 3", "Score 4", "Score 5", "Average", "Status")
records_list = ttk.Treeview(records_frame, columns=columns, show="headings")

for col in columns:
    records_list.heading(col, text=col)
    records_list.column(col, anchor="center", width=90)

scroll_y = ttk.Scrollbar(records_frame, orient="vertical", command=records_list.yview)
scroll_x = ttk.Scrollbar(records_frame, orient="horizontal", command=records_list.xview)
records_list.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

records_list.grid(row=0, column=0, sticky="nsew")
scroll_y.grid(row=0, column=1, sticky="ns")
scroll_x.grid(row=1, column=0, sticky="ew")

# Buttons
ctk.CTkButton(mw, text="Calculate Average", command=calculate_average_event).grid(row=7, column=0, padx=10, pady=10, columnspan=2)
ctk.CTkButton(mw, text="Save Record", command=save_record).grid(row=7, column=3, padx=10, pady=10, sticky="e")
ctk.CTkButton(mw, text="Clear Entries", command=clear_entries).grid(row=8, column=0, padx=10, pady=10, columnspan=2)
ctk.CTkButton(mw, text="Clear Table", command=clear_table).grid(row=8, column=3, padx=10, pady=10, sticky="w")
ctk.CTkButton(mw, text="Refresh Table", command=display_all_records).grid(row=7, column=3, padx=10, pady=10, sticky="w")
ctk.CTkButton(mw, text="Exit", command=quit).grid(row=8, column=3, padx=10, pady=10, sticky="e")

# Load existing records on start
display_all_records()
mw.mainloop()